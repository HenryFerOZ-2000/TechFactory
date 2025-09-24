import calendar
from random import randint
from datetime import date, timedelta, datetime
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.timezone import now
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from django.utils import timezone
from django.utils import timezone  # ya lo tienes importado arriba
from base.models import Persona
from django.contrib.admin.views.decorators import staff_member_required
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils import timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils import timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils import timezone
from .models import Impresora, Reserva, HOURS_RANGE
from .forms import PublicReservationForm
from django.shortcuts import redirect
from django.urls import reverse
from types import SimpleNamespace
from .models import Impresora, Reserva, LabReserva, HOURS_RANGE
from .forms import PublicReservationForm, PublicLabReservationForm
from datetime import date, timedelta, datetime
import calendar
from io import BytesIO
from django import template
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.db.models import Case, When, IntegerField, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from base.models import Persona
from .forms import PublicReservationForm, PublicLabReservationForm
from .models import Impresora, Reserva, LabReserva, HOURS_RANGE


LAB_NAME = "Laboratorio Tech Factory"
LAB_CAPACITY = 15

register = template.Library()
from django.contrib.admin.views.decorators import staff_member_required
from base.models import Persona


@staff_member_required
def lista_penalizados(request):
    """Lista personas penalizadas: incluye quienes tengan penalización en impresoras o laboratorio."""
    now_ = timezone.now()
    personas = (Persona.objects
                .filter(Q(penalizado_impresoras_hasta__gt=now_) | Q(penalizado_lab_hasta__gt=now_))
                .order_by('-penalizado_impresoras_hasta', '-penalizado_lab_hasta', 'nombre'))
    return render(request, "impresoras/penalizados.html", {"personas": personas})

# ---------------------------------------------------------------------
# Utilidades comunes
# ---------------------------------------------------------------------
# --- HELPERS PERSONA ---

def _find_persona(cedula: str = "", nombre: str = ""):
    """
    Devuelve Persona si existe (prioriza cédula, luego nombre) o None si no hay.
    No crea nada.
    """
    from base.models import Persona
    ced = (cedula or "").strip()
    nom = (nombre or "").strip()
    p = None
    if ced:
        p = Persona.objects.filter(cedula=ced).first()
    if not p and nom:
        p = Persona.objects.filter(nombre__iexact=nom).first()
    return p


def get_or_create_persona(nombre: str, cedula: str, celular: str, carrera: str):
    """
    Busca por cédula o nombre y, si no existe, crea un Persona con los datos de la reserva.
    Si existe, actualiza nombre/celular/carrera cuando vengan informados.
    Devuelve la instancia de Persona.
    """
    from base.models import Persona

    p = _find_persona(cedula, nombre)

    if p:
        changed = False
        nom = (nombre or "").strip()
        cel = (celular or "").strip()
        car = (carrera or "") or p.carrera

        if nom and p.nombre != nom:
            p.nombre = nom; changed = True
        if cel and p.celular != cel:
            p.celular = cel; changed = True
        if car and p.carrera != car:
            p.carrera = car; changed = True

        if changed:
            p.save(update_fields=["nombre", "celular", "carrera"])
        return p

    # crear una cédula auto si no vino
    ced = (cedula or "").strip()
    if not ced:
        ced = f"AUTO-{timezone.now().strftime('%Y%m%d')}-{randint(1000, 9999)}"

    return Persona.objects.create(
        nombre=(nombre or "Sin nombre").strip(),
        cedula=ced,
        celular=(celular or "").strip(),
        carrera=carrera or Persona.CARRERAS[0][0],
    )


def _next(request, fallback_name: str):
    """
    Adonde volver luego de una acción:
    - POST.next (del modal)
    - GET.next (enlaces admin)
    - Referer del navegador
    - reverse(fallback_name) como último recurso
    """
    return (
        request.POST.get("next")
        or request.GET.get("next")
        or request.META.get("HTTP_REFERER")
        or reverse(fallback_name)
    )

def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())

def generate_week_days(pivot: date):
    monday = monday_of_week(pivot)
    return [monday + timedelta(days=i) for i in range(6)]  # Lun a sabado

def reservations_map(impresora, dias):
    qs = Reserva.objects.filter(impresora=impresora, fecha__in=dias).exclude(estado='cancelado')
    return {(r.fecha, r.hora): r for r in qs}

def list_impresoras():
    # 0 = impresoras, 1 = laboratorio → así el lab queda al final
    return list(
        Impresora.objects
        .annotate(
            is_lab=Case(
                When(nombre__iexact=LAB_NAME, then=1),
                default=0,
                output_field=IntegerField(),
            )
        )
        .order_by('is_lab', 'nombre')
    )

# ---------------------------------------------------------------------
# Mapas de reservas LAB
# ---------------------------------------------------------------------
def lab_reservations_map(dias):
    """
    Devuelve un dict {(fecha, hora): {'count': N, 'items': [LabReserva,...]}}
    Solo cuenta reservas no canceladas.
    """
    qs = LabReserva.objects.filter(fecha__in=dias).exclude(estado='cancelado')
    bucket = {}
    for r in qs:
        key = (r.fecha, r.hora)
        bucket.setdefault(key, []).append(r)
    return {k: {'count': len(v), 'items': v} for k, v in bucket.items()}

# ---------------------------------------------------------------------
# Helpers de BUFFER y Penalización
# ---------------------------------------------------------------------
def crear_buffer_si_libre(reserva_normal: Reserva):
    """Crea la hora siguiente como BUFFER si está libre (mismo día/impresora)."""
    if reserva_normal.tipo != "NORMAL":
        return None
    siguiente = reserva_normal.hora + 1
    if siguiente not in HOURS_RANGE:
        return None
    ya_existe = Reserva.objects.filter(
        impresora=reserva_normal.impresora,
        fecha=reserva_normal.fecha,
        hora=siguiente
    ).exists()
    if ya_existe:
        return None
    return Reserva.objects.create(
        impresora=reserva_normal.impresora,
        fecha=reserva_normal.fecha,
        hora=siguiente,
        estudiante_nombre=reserva_normal.estudiante_nombre,
        estudiante_cedula=reserva_normal.estudiante_cedula,
        estudiante_celular=reserva_normal.estudiante_celular,
        estudiante_carrera=reserva_normal.estudiante_carrera,
        estado='reservado',
        tipo='BUFFER',
        parent=reserva_normal,
        tecnico_observaciones="Bloque de margen automático (BUFFER 1h)."
    )

def eliminar_buffer_de(reserva_normal: Reserva):
    Reserva.objects.filter(parent=reserva_normal, tipo='BUFFER').delete()

# impresoras/views.py (o donde tengas el helper)
from base.models import Persona

def persona_de_reserva(obj) -> Persona | None:
    """
    Obtiene o crea Persona a partir de una Reserva, LabReserva o un objeto compatible.
    Tolera que falten atributos; prioriza cédula y luego nombre.
    Además, si la Persona existe pero tiene campos vacíos, los completa.
    """
    # Toma atributos de forma segura (sin reventar si no existen)
    ced = (getattr(obj, "estudiante_cedula", "") or getattr(obj, "cedula", "") or "").strip()
    nom = (getattr(obj, "estudiante_nombre", "") or getattr(obj, "nombre", "") or "").strip()
    cel = (getattr(obj, "estudiante_celular", "") or getattr(obj, "celular", "") or "").strip()
    car = (getattr(obj, "estudiante_carrera", "") or getattr(obj, "carrera", "") or "") \
          or Persona.CARRERAS[0][0]

    if not ced and not nom:
        return None  # no hay base para identificar/crear

    # Busca por cédula si hay; si no, por nombre
    if ced:
        p, _ = Persona.objects.get_or_create(
            cedula=ced,
            defaults={"nombre": nom or ced, "celular": cel, "carrera": car},
        )
    else:
        p, _ = Persona.objects.get_or_create(
            nombre=nom,
            defaults={"cedula": "", "celular": cel, "carrera": car},
        )

    # Completa campos faltantes si la Persona ya existía
    changed = False
    if nom and not (p.nombre or "").strip():
        p.nombre = nom; changed = True
    if cel and not (p.celular or "").strip():
        p.celular = cel; changed = True
    if car and not (p.carrera or "").strip():
        p.carrera = car; changed = True
    if ced and not (p.cedula or "").strip():
        p.cedula = ced; changed = True
    if changed:
        p.save()

    return p


def persona_de_lab(lab: LabReserva) -> Persona | None:
    """Obtiene o crea Persona a partir de una reserva de laboratorio."""
    ced = (lab.estudiante_cedula or "").strip()
    nom = (lab.estudiante_nombre or "").strip()
    cel = (lab.estudiante_celular or "").strip()
    car = (lab.estudiante_carrera or "") or Persona.CARRERAS[0][0]

    if ced:
        p, _ = Persona.objects.get_or_create(
            cedula=ced,
            defaults={"nombre": nom or ced, "celular": cel, "carrera": car},
        )
        return p
    if nom:
        p, _ = Persona.objects.get_or_create(
            nombre=nom,
            defaults={"cedula": "", "celular": cel, "carrera": car},
        )
        return p
    return None

# ---------------------------------------------------------------------
# Build context
# ---------------------------------------------------------------------
def _build_context(pivot_date: date, admin: bool):
    week_days = generate_week_days(pivot_date)
    impresoras = list_impresoras()

    tab_data = []
    for imp in impresoras:
        # ----- LAB -----
        if imp.nombre.strip().lower() == LAB_NAME.lower():
            # Mapa base: {(fecha, hora): {'count': N, 'items': [LabReserva,...]}}
            base = lab_reservations_map(week_days)

            # Normalizamos TODAS las celdas para la semana y calculamos all_used
            norm_mapa = {}
            for d in week_days:
                for h in HOURS_RANGE:
                    key = (d, h)
                    data = base.get(key, {'count': 0, 'items': []})
                    items = data.get('items', [])
                    count = int(data.get('count', 0) or 0)
                    all_used = (count > 0) and all(getattr(x, "estado", "") == "usado" for x in items)

                    norm_mapa[key] = {
                        'count': count,
                        'items': items,
                        'all_used': all_used,
                    }

            mapa = norm_mapa
            is_lab = True

        # ----- IMPRESORAS -----
        else:
            mapa = reservations_map(imp, week_days)
            is_lab = False

        tab_data.append({"imp": imp, "mapa": mapa, "is_lab": is_lab})

    return {
        "week_days": week_days,
        "hours": HOURS_RANGE,
        "tab_data": tab_data,
        "is_admin": admin,
        "form": PublicReservationForm(),
        "lab_name": LAB_NAME,
        "lab_capacity": LAB_CAPACITY,
    }

# ---------------------------------------------------------------------
# Público
# ---------------------------------------------------------------------
def calendario_publico(request):
    week_str = request.GET.get("week")
    pivot = datetime.strptime(week_str, "%Y-%m-%d").date() if week_str else now().date()
    ctx = _build_context(pivot, admin=False)
    return render(request, "impresoras/calendario.html", ctx)

def crear_reserva_lab(request):
    if request.method != "POST":
        next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('impresoras:calendario')
        return redirect(next_url)

    form = PublicLabReservationForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Datos inválidos.")
        return redirect(request.POST.get('next') or "impresoras:calendario")

    f = form.cleaned_data["fecha"]
    h = form.cleaned_data["hora"]

    today = timezone.localdate()
    now_hour = timezone.localtime().hour
    if f < today:
        messages.error(request, "No puedes reservar una fecha pasada.")
        return redirect(request.POST.get('next') or "impresoras:calendario")
    if f == today and h < now_hour:
        messages.error(request, "No puedes reservar una hora que ya pasó.")
        return redirect(request.POST.get('next') or "impresoras:calendario")

    if f.weekday() > 4 or h not in HOURS_RANGE:
        messages.error(request, "Horario fuera de rango (Lun–Vie 08:00–21:00).")
        return redirect(request.POST.get('next') or "impresoras:calendario")

    count = LabReserva.objects.filter(fecha=f, hora=h).exclude(estado='cancelado').count()
    if count >= LAB_CAPACITY:
        messages.error(request, "No hay cupos disponibles en esa franja.")
        return redirect(request.POST.get('next') or "impresoras:calendario")

    # ------------ Checar penalización y crear Persona ------------
    ced = form.cleaned_data.get("estudiante_cedula") or ""
    nom = form.cleaned_data.get("estudiante_nombre") or ""
    persona_existente = _find_persona(ced, nom)
    if persona_existente and persona_existente.penalizado_lab_hasta and persona_existente.penalizado_lab_hasta > timezone.now():
        messages.error(request, "No puedes reservar: estás penalizado temporalmente en laboratorio.")
        return redirect(request.POST.get('next') or "impresoras:calendario")

    persona = get_or_create_persona(
        nombre=nom,
        cedula=ced,
        celular=form.cleaned_data.get("estudiante_celular") or "",
        carrera=form.cleaned_data.get("estudiante_carrera") or "",
    )
    # -------------------------------------------------------------

    # Actividad (nuevo campo)
    actividad = (form.cleaned_data.get("actividad") or "").strip()

    LabReserva.objects.create(
        fecha=f, hora=h,
        estudiante_nombre=form.cleaned_data["estudiante_nombre"],
        estudiante_cedula=form.cleaned_data["estudiante_cedula"],
        estudiante_celular=form.cleaned_data["estudiante_celular"],
        estudiante_carrera=form.cleaned_data["estudiante_carrera"],
        actividad=actividad,
        estado='reservado'
    )
    messages.success(request, "Reserva de laboratorio creada correctamente.")
    return redirect(request.POST.get('next') or "impresoras:calendario")

def crear_reserva(request):
    if request.method != "POST":
        return redirect(_next(request, "impresoras:calendario"))

    form = PublicReservationForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Datos inválidos.")
        return redirect(_next(request, "impresoras:calendario"))

    imp = get_object_or_404(Impresora, id=form.cleaned_data["impresora_id"])
    f = form.cleaned_data["fecha"]
    h = form.cleaned_data["hora"]

    today = timezone.localdate()
    now_hour = timezone.localtime().hour
    if f < today:
        messages.error(request, "No puedes reservar una fecha pasada.")
        next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('impresoras:calendario')
        return redirect(next_url)
    if f == today and h < now_hour:
        messages.error(request, "No puedes reservar una hora que ya pasó.")
        next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('impresoras:calendario')
        return redirect(next_url)

    # Lun–Vie y rango de horas permitidas
    if f.weekday() > 5 or h not in HOURS_RANGE:
        messages.error(request, "Horario fuera de rango (Lun–Vie 08:00–21:00).")
        return redirect(_next(request, "impresoras:calendario"))

    # La hora solicitada ya está tomada
    if Reserva.objects.filter(impresora=imp, fecha=f, hora=h).exists():
        messages.error(request, "Esa hora ya está reservada.")
        return redirect(_next(request, "impresoras:calendario"))

    # ------------ NUEVO: checar penalización si ya existía la persona ------------
    ced = form.cleaned_data.get("estudiante_cedula") or ""
    nom = form.cleaned_data.get("estudiante_nombre") or ""
    persona_existente = _find_persona(ced, nom)
    if persona_existente and persona_existente.penalizado_impresoras_hasta and persona_existente.penalizado_impresoras_hasta > timezone.now():
        messages.error(request, "No puedes reservar: estás penalizado temporalmente en impresoras.")
        return redirect(_next(request, "impresoras:calendario"))

    # Crear/actualizar Persona en este punto (si no existía, la crea)
    persona = get_or_create_persona(
        nombre=nom,
        cedula=ced,
        celular=form.cleaned_data.get("estudiante_celular") or "",
        carrera=form.cleaned_data.get("estudiante_carrera") or "",
    )
    # ------------ FIN NUEVO ------------------------------------------------------

    # --- Crear reserva NORMAL ---
    reserva = Reserva.objects.create(
        impresora=imp,
        fecha=f,
        hora=h,
        estudiante_nombre=form.cleaned_data["estudiante_nombre"],
        estudiante_cedula=form.cleaned_data["estudiante_cedula"],
        estudiante_celular=form.cleaned_data["estudiante_celular"],
        estudiante_carrera=form.cleaned_data["estudiante_carrera"],
        estado='reservado',
        tipo='NORMAL'
    )

    # --- Crear BUFFER 1h si está libre ---
    creado_buffer = False
    siguiente = h + 1
    if siguiente in HOURS_RANGE:
        ocupado = Reserva.objects.filter(
            impresora=imp,
            fecha=f,
            hora=siguiente
        ).exists()
        if not ocupado:
            Reserva.objects.create(
                impresora=imp,
                fecha=f,
                hora=siguiente,
                estudiante_nombre=reserva.estudiante_nombre,
                estudiante_cedula=reserva.estudiante_cedula,
                estudiante_celular=reserva.estudiante_celular,
                estudiante_carrera=reserva.estudiante_carrera,
                estado='reservado',
                tipo='BUFFER',
                parent=reserva,
                tecnico_observaciones="Bloque de margen automático (BUFFER 1h)."
            )
            creado_buffer = True

    if creado_buffer:
        messages.success(request, "Reserva creada correctamente. Se añadió 1 hora de margen.")
    else:
        messages.success(request, "Reserva creada correctamente.")

    return redirect(_next(request, "impresoras:calendario"))

# ---------------------------------------------------------------------
# Penalizaciones y buffer (admin)
# ---------------------------------------------------------------------
@staff_member_required
def penalizar_reserva(request, reserva_id):
    from datetime import timedelta
    reserva = get_object_or_404(Reserva, id=reserva_id)
    persona = persona_de_reserva(reserva)
    days = int(request.POST.get("days", 3))

    if not persona:
        messages.error(request, "No se encontró la Persona asociada a esta reserva.")
        return redirect(_next(request, "impresoras:admin_reservas"))

    # Penalización a la persona (IMPRESORAS)
    persona.no_show_count = (persona.no_show_count or 0) + 1
    persona.penalizado_impresoras_hasta = timezone.now() + timedelta(days=days)
    persona.save(update_fields=["no_show_count", "penalizado_impresoras_hasta"])

    # La reserva queda como PENALIZADO y se elimina su margen
    reserva.estado = 'penalizado'
    reserva.save(update_fields=['estado'])
    eliminar_buffer_de(reserva)

    messages.success(request, f"{persona.nombre} penalizado por {days} días (impresoras).")
    return redirect(_next(request, "impresoras:admin_reservas"))


@staff_member_required
def penalizar_lab(request, lab_id):
    """Penaliza a un alumno desde una reserva del laboratorio (no borra la reserva)."""
    lab = get_object_or_404(LabReserva, id=lab_id)
    days = int(request.POST.get("days", 3))
    next_url = _next(request, "impresoras:calendario")

    persona = persona_de_lab(lab)
    if not persona:
        messages.error(request, "No se pudo identificar a la persona.")
        return redirect(next_url)

    # Penalización a la persona (LABORATORIO)
    persona.no_show_count = (persona.no_show_count or 0) + 1
    persona.penalizado_lab_hasta = timezone.now() + timedelta(days=days)
    persona.save(update_fields=["no_show_count", "penalizado_lab_hasta"])

    messages.success(request, f"Penalización aplicada a {persona.nombre} (laboratorio).")
    return redirect(next_url)

@staff_member_required
def liberar_buffer(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id)
    eliminar_buffer_de(reserva)
    messages.success(request, "Margen liberado.")
    return redirect(_next(request, "impresoras:calendario"))

# ---------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------
@login_required
@login_required
def exportar_excel_mes(request):
    from collections import defaultdict

    month_str = request.GET.get("month")           # "YYYY-MM"
    impresora_id = request.GET.get("impresora")    # puede venir vacío o ser un id

    if not month_str:
        messages.error(request, "Selecciona un mes (YYYY-MM).")
        return redirect("impresoras:calendario")

    try:
        y, m = map(int, month_str.split("-"))
        first = date(y, m, 1)
        last  = date(y, m, calendar.monthrange(y, m)[1])
    except Exception:
        messages.error(request, "Formato de mes inválido (usa YYYY-MM).")
        return redirect("impresoras:calendario")

    # identifica ids del LAB por nombre
    lab_ids = set(map(str, Impresora.objects
                      .filter(nombre__iexact=LAB_NAME)  # más estricto y claro
                      .values_list("id", flat=True)))

    # ========= Construcción de datos =========
    rows_imp = []   # impresoras normales (hoja "Reservas")
    rows_lab = []   # detalle laboratorio (hoja "Detalle Lab" + ordenación técnica)

    # ----- Caso 1: TODAS -----
    if not impresora_id:
        qs_imp = (Reserva.objects
                  .filter(fecha__range=(first, last))
                  .select_related("impresora")
                  .order_by("impresora__nombre", "fecha", "hora"))
        for r in qs_imp:
            creado_local = timezone.localtime(r.creado_en).replace(tzinfo=None)
            rows_imp.append((
                r.impresora.nombre, r.fecha,
                f"{r.hora:02d}:00 - {r.hora+1:02d}:00",
                r.get_estado_display(), r.estudiante_nombre,
                r.estudiante_cedula, r.estudiante_celular, r.estudiante_carrera,
                (getattr(r, "tecnico_observaciones", "") or "").replace("\r\n", " ").replace("\n", " ").strip(),
                creado_local,
            ))

        qs_lab = (LabReserva.objects
                  .filter(fecha__range=(first, last))
                  .order_by("fecha", "hora", "estado", "estudiante_nombre"))
        for r in qs_lab:
            creado_local = timezone.localtime(r.creado_en).replace(tzinfo=None)
            rows_lab.append((
                LAB_NAME, r.fecha,
                f"{r.hora:02d}:00 - {r.hora+1:02d}:00",
                r.get_estado_display(), r.estudiante_nombre,
                r.estudiante_cedula, r.estudiante_celular, r.estudiante_carrera,
                (getattr(r, "tecnico_observaciones", "") or "").strip().replace("\r\n", " ").replace("\n", " "),
                creado_local,
                r.hora,          # <- índice 10 (técnico para ordenar)
                r.estado,        # <- índice 11 (técnico para ordenar)
            ))

    # ----- Caso 2: SOLO LAB -----
    elif impresora_id in lab_ids:
        qs_lab = (LabReserva.objects
                  .filter(fecha__range=(first, last))
                  .order_by("fecha", "hora", "estado", "estudiante_nombre"))
        for r in qs_lab:
            creado_local = timezone.localtime(r.creado_en).replace(tzinfo=None)
            rows_lab.append((
                LAB_NAME, r.fecha,
                f"{r.hora:02d}:00 - {r.hora+1:02d}:00",
                r.get_estado_display(), r.estudiante_nombre,
                r.estudiante_cedula, r.estudiante_celular, r.estudiante_carrera,
                (getattr(r, "tecnico_observaciones", "") or "").strip().replace("\r\n", " ").replace("\n", " "),
                creado_local,
                r.hora,          # 10
                r.estado,        # 11
            ))

        # en este caso, también exporta si hubiera reservas en una impresora con ese id (por coherencia)
        qs_imp_lab = (Reserva.objects
                      .filter(fecha__range=(first, last), impresora_id=impresora_id)
                      .select_related("impresora")
                      .order_by("fecha", "hora"))
        for r in qs_imp_lab:
            creado_local = timezone.localtime(r.creado_en).replace(tzinfo=None)
            rows_imp.append((
                r.impresora.nombre, r.fecha,
                f"{r.hora:02d}:00 - {r.hora+1:02d}:00",
                r.get_estado_display(), r.estudiante_nombre,
                r.estudiante_cedula, r.estudiante_celular, r.estudiante_carrera,
                (getattr(r, "tecnico_observaciones", "") or "").replace("\r\n", " ").replace("\n", " ").strip(),
                creado_local,
            ))

    # ----- Caso 3: SOLO UNA IMPRESORA NORMAL -----
    else:
        qs_imp = (Reserva.objects
                  .filter(fecha__range=(first, last), impresora_id=impresora_id)
                  .select_related("impresora")
                  .order_by("impresora__nombre", "fecha", "hora"))
        for r in qs_imp:
            creado_local = timezone.localtime(r.creado_en).replace(tzinfo=None)
            rows_imp.append((
                r.impresora.nombre, r.fecha,
                f"{r.hora:02d}:00 - {r.hora+1:02d}:00",
                r.get_estado_display(), r.estudiante_nombre,
                r.estudiante_cedula, r.estudiante_celular, r.estudiante_carrera,
                (getattr(r, "tecnico_observaciones", "") or "").replace("\r\n", " ").replace("\n", " ").strip(),
                creado_local,
            ))

    # ========= Workbook =========
    wb = Workbook()
    ws_reservas = wb.active
    ws_reservas.title = "Reservas"

    headers = [
        "Recurso", "Fecha", "Hora", "Estado",
        "Nombre", "Cédula", "Celular", "Carrera",
        "Observaciones", "Creado",
    ]
    ws_reservas.append(headers)
    for row in rows_imp:
        ws_reservas.append(row)

    # Formato fecha/hora de "Creado"
    for cell in ws_reservas["J"][1:]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for c in ws_reservas[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="DDDDDD")
    for row in ws_reservas.iter_rows(min_row=1, max_row=ws_reservas.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Auto ancho
    for col_cells in ws_reservas.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws_reservas.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)

    # Table
    if ws_reservas.max_row > 1:
        table = Table(displayName="TablaReservas", ref=f"A1:J{ws_reservas.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
        )
        ws_reservas.add_table(table)
    else:
        ws_reservas["A2"] = "Sin reservas en el período/selector elegido."
        ws_reservas.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws_reservas["A2"].alignment = Alignment(horizontal="center")
        ws_reservas["A2"].font = Font(italic=True, color="666666")

    # ========= Hojas del LAB (si corresponde) =========
    include_lab_sheets = (not impresora_id) or (impresora_id in lab_ids)
    if include_lab_sheets:
        # ---- Hoja 2: Detalle Lab ----
        ws_lab_det = wb.create_sheet("Detalle Lab")
        lab_headers = [
            "Recurso", "Fecha", "Hora", "Estado",
            "Nombre", "Cédula", "Celular", "Carrera",
            "Observaciones", "Creado",
        ]
        ws_lab_det.append(lab_headers)

        # Orden: Fecha, Hora (técnico 10), Estado bruto (técnico 11: 'usado' primero)
        estado_order = {"usado": 0, "reservado": 1}
        rows_lab_sorted = sorted(
            rows_lab,
            key=lambda x: (x[1], x[10], estado_order.get((x[11] or "").lower(), 99), x[4] or "")
        )
        for r in rows_lab_sorted:
            ws_lab_det.append(r[:10])  # sólo columnas visibles

        # Formato "Creado"
        for cell in ws_lab_det["J"][1:]:
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"

        # Estilo
        for c in ws_lab_det[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws_lab_det.iter_rows(min_row=1, max_row=ws_lab_det.max_row, min_col=1, max_col=len(lab_headers)):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for col_cells in ws_lab_det.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws_lab_det.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)

        if ws_lab_det.max_row > 1:
            t2 = Table(displayName="TablaDetalleLab", ref=f"A1:J{ws_lab_det.max_row}")
            t2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws_lab_det.add_table(t2)
        else:
            ws_lab_det["A2"] = "Sin reservas de laboratorio en el mes."
            ws_lab_det.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(lab_headers))
            ws_lab_det["A2"].alignment = Alignment(horizontal="center")
            ws_lab_det["A2"].font = Font(italic=True, color="666666")

        # ---- Hoja 3: Resumen Lab ----
        resumen = defaultdict(lambda: {"total": 0, "usados": 0, "reservados": 0})
        for r in rows_lab:
            fecha = r[1]
            hora_num = r[10]  # columna técnica
            estado_raw = (r[11] or "").lower()  # columna técnica
            resumen[(fecha, hora_num)]["total"] += 1
            if estado_raw == "usado":
                resumen[(fecha, hora_num)]["usados"] += 1
            elif estado_raw == "reservado":
                resumen[(fecha, hora_num)]["reservados"] += 1

        ws_lab_sum = wb.create_sheet("Resumen Lab")
        sum_headers = ["Fecha", "Hora", "Total", "Usados", "Reservados", "Cupo", "%Uso (Usados/Cupo)"]
        ws_lab_sum.append(sum_headers)

        for (fecha, hora_num) in sorted(resumen.keys()):
            tot = resumen[(fecha, hora_num)]["total"]
            usados = resumen[(fecha, hora_num)]["usados"]
            reservados = resumen[(fecha, hora_num)]["reservados"]
            cupo = LAB_CAPACITY
            pct = usados / cupo if cupo else 0
            ws_lab_sum.append([
                fecha,
                f"{hora_num:02d}:00 - {hora_num+1:02d}:00",
                tot, usados, reservados, cupo, pct
            ])

        # Estilo encabezado
        for c in ws_lab_sum[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Formato %
        for cell in ws_lab_sum["G"][1:]:
            if isinstance(cell.value, float):
                cell.number_format = "0.00%"

        # Bordes y anchos
        for row in ws_lab_sum.iter_rows(min_row=1, max_row=ws_lab_sum.max_row, min_col=1, max_col=len(sum_headers)):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for col_cells in ws_lab_sum.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws_lab_sum.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 32)

        if ws_lab_sum.max_row > 1:
            t3 = Table(displayName="TablaResumenLab", ref=f"A1:G{ws_lab_sum.max_row}")
            t3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
            ws_lab_sum.add_table(t3)
        else:
            ws_lab_sum["A2"] = "Sin datos para el resumen de laboratorio."
            ws_lab_sum.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(sum_headers))
            ws_lab_sum["A2"].alignment = Alignment(horizontal="center")
            ws_lab_sum["A2"].font = Font(italic=True, color="666666")

        # ---- Hoja 4 (opcional): Asistencia Lab (solo USADOS) ----
        ws_lab_used = wb.create_sheet("Asistencia Lab (Usados)")
        used_headers = ["Fecha", "Hora", "Nombre", "Cédula", "Celular", "Carrera", "Observaciones", "Creado"]
        ws_lab_used.append(used_headers)

        rows_lab_used = [r for r in rows_lab if (r[11] or "").lower() == "usado"]
        rows_lab_used = sorted(rows_lab_used, key=lambda r: (r[1], r[10], r[4] or ""))
        for r in rows_lab_used:
            ws_lab_used.append([
                r[1], r[2],  # Fecha, Hora (texto)
                r[4], r[5], r[6], r[7],  # Nombre, Cédula, Celular, Carrera
                r[8], r[9],  # Observaciones, Creado
            ])

        for c in ws_lab_used[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws_lab_used.iter_rows(min_row=1, max_row=ws_lab_used.max_row, min_col=1, max_col=len(used_headers)):
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for col_cells in ws_lab_used.columns:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws_lab_used.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)
        if ws_lab_used.max_row > 1:
            t4 = Table(displayName="TablaAsistenciaLab", ref=f"A1:H{ws_lab_used.max_row}")
            t4.tableStyleInfo = TableStyleInfo(name="TableStyleMedium7", showRowStripes=True)
            ws_lab_used.add_table(t4)
        else:
            ws_lab_used["A2"] = "Sin asistencias registradas (estado 'usado') en el mes."
            ws_lab_used.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(used_headers))
            ws_lab_used["A2"].alignment = Alignment(horizontal="center")
            ws_lab_used["A2"].font = Font(italic=True, color="666666")

    # ========= Respuesta =========
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="reservas_{month_str}.xlsx"'
    return resp


# ---------------------------------------------------------------------
# ADMIN (con login)
# ---------------------------------------------------------------------
@login_required
def calendario_admin(request):
    week_str = request.GET.get("week")
    pivot = datetime.strptime(week_str, "%Y-%m-%d").date() if week_str else now().date()
    ctx = _build_context(pivot, admin=True)
    return render(request, "impresoras/calendario.html", ctx)



@staff_member_required
def penalizar_lab(request, lab_id):
    lab = get_object_or_404(LabReserva, id=lab_id)
    # buscar Persona por cédula y, si no, por nombre:
    persona = None
    if lab.estudiante_cedula:
        persona = Persona.objects.filter(cedula=lab.estudiante_cedula).first()
    if not persona and lab.estudiante_nombre:
        persona = Persona.objects.filter(nombre=lab.estudiante_nombre).first()

    if not persona:
        messages.error(request, "No se encontró la Persona asociada a esta reserva.")
        return redirect(_next(request, "impresoras:admin_reservas"))

    days = int(request.POST.get("days", 3))
    # Penalización específica de LABORATORIO
    persona.no_show_count = (getattr(persona, "no_show_count", 0) or 0) + 1
    persona.penalizado_lab_hasta = timezone.now() + timedelta(days=days)
    persona.save(update_fields=["no_show_count", "penalizado_lab_hasta"])

    messages.success(request, f"Penalización aplicada a {persona.nombre}.")
    return redirect(_next(request, "impresoras:admin_reservas"))


@login_required
def admin_marcar_usado(request, pk):
    r = get_object_or_404(Reserva, pk=pk)

    if request.method == "POST":
        obs = request.POST.get("observaciones", "").strip()
        r.tecnico_observaciones = obs
        r.estado = "usado"
        r.save()
        messages.success(request, "Reserva marcada como usada.")
    else:
        if r.estado != "usado":
            r.estado = "usado"
            r.save()
            messages.success(request, "Reserva marcada como usada.")
        else:
            messages.info(request, "Esta reserva ya estaba marcada como usada.")

    next_url = _next(request, "impresoras:calendario")
    return redirect(next_url)

@login_required
def admin_cancelar_reserva(request, pk):
    r = get_object_or_404(Reserva, pk=pk)

    # Si es NORMAL, borra su buffer antes:
    if r.tipo == 'NORMAL':
        Reserva.objects.filter(parent=r, tipo='BUFFER').delete()

    r.delete()  # elimina la reserva (y libera el UNIQUE)
    messages.success(request, "Reserva cancelada.")
    return redirect(_next(request, "impresoras:admin_reservas"))


@login_required
def admin_cancelar_lab(request, pk):
    r = get_object_or_404(LabReserva, pk=pk)
    r.delete()
    messages.success(request, "Reserva cancelada. La franja quedó libre.")
    return redirect(_next(request, "impresoras:calendario"))

@login_required
def admin_marcar_usado_lab(request, pk):
    r = get_object_or_404(LabReserva, pk=pk)
    if request.method == "POST":
        obs = request.POST.get("observaciones", "").strip()
        r.tecnico_observaciones = obs
        r.estado = "usado"
        r.save()
        messages.success(request, "Reserva marcada como usada.")
    else:
        if r.estado != "usado":
            r.estado = "usado"
            r.save()
            messages.success(request, "Reserva marcada como usada.")
        else:
            messages.info(request, "Esta reserva ya estaba marcada como usada.")
    return redirect(_next(request, "impresoras:calendario"))

# ---------------------------------------------------------------------
# LISTA DE PENALIZADOS (Admin)
# ---------------------------------------------------------------------
@staff_member_required
def lista_penalizados_imp(request):
    now_ = timezone.now()
    personas = (Persona.objects
                .filter(penalizado_impresoras_hasta__isnull=False,
                        penalizado_impresoras_hasta__gt=now_)
                .order_by("-penalizado_impresoras_hasta", "nombre"))
    ctx = {"titulo": "Penalizados • Impresoras", "personas": personas, "tipo": "impresoras"}
    return render(request, "impresoras/lista_penalizados_tipo.html", ctx)

@staff_member_required
def lista_penalizados_lab(request):
    now_ = timezone.now()
    personas = (Persona.objects
                .filter(penalizado_lab_hasta__isnull=False,
                        penalizado_lab_hasta__gt=now_)
                .order_by("-penalizado_lab_hasta", "nombre"))
    ctx = {"titulo": "Penalizados • Laboratorio", "personas": personas, "tipo": "lab"}
    return render(request, "impresoras/lista_penalizados_tipo.html", ctx)



@staff_member_required
def penalizar_lab(request, lab_id):
    from datetime import timedelta
    r = get_object_or_404(LabReserva, id=lab_id)
    persona = persona_de_reserva(SimpleNamespace(
        estudiante_cedula=r.estudiante_cedula,
        estudiante_nombre=r.estudiante_nombre
    ))
    days = int(request.POST.get("days", 3))

    if not persona:
        messages.error(request, "No se encontró la Persona asociada a esta reserva.")
        return redirect(_next(request, "impresoras:admin_reservas"))

    # Penalización específica de LABORATORIO (segunda definición mantenida, ajustada)
    persona.no_show_count = (persona.no_show_count or 0) + 1
    persona.penalizado_lab_hasta = timezone.now() + timedelta(days=days)
    persona.save(update_fields=["no_show_count", "penalizado_lab_hasta"])

    r.estado = 'penalizado'
    r.save(update_fields=['estado'])

    messages.success(request, f"{persona.nombre} penalizado por {days} días (laboratorio).")
    return redirect(_next(request, "impresoras:admin_reservas"))



@staff_member_required
def lista_penalizados_impresoras(request):
    now_ = timezone.now()
    personas = (Persona.objects
                .filter(penalizado_impresoras_hasta__isnull=False,
                        penalizado_impresoras_hasta__gt=now_)
                .order_by('-penalizado_impresoras_hasta', 'nombre'))
    return render(request, "impresoras/penalizados.html",
                  {"personas": personas, "titulo": "Penalizados (impresoras)"})


@staff_member_required
def lista_penalizados_laboratorio(request):
    now_ = timezone.now()
    personas = (Persona.objects
                .filter(penalizado_lab_hasta__isnull=False,
                        penalizado_lab_hasta__gt=now_)
                .order_by('-penalizado_lab_hasta', 'nombre'))
    return render(request, "impresoras/penalizados.html",
                  {"personas": personas, "titulo": "Penalizados (laboratorio)"})



@staff_member_required
def penalizados_tabs(request):
    """
    Página con pestañas: 'Impresoras' y 'Laboratorio'.
    Muestra únicamente personas con penalización vigente en cada categoría.
    Usa un pequeño margen para evitar edge-cases de TZ/redondeo.
    """
    now_ = timezone.now()
    safety_margin = timedelta(minutes=1)

    imp_pen = (Persona.objects
               .filter(penalizado_impresoras_hasta__isnull=False)
               .filter(penalizado_impresoras_hasta__gte=now_ - safety_margin)
               .order_by('-penalizado_impresoras_hasta', 'nombre'))

    lab_pen = (Persona.objects
               .filter(penalizado_lab_hasta__isnull=False)
               .filter(penalizado_lab_hasta__gte=now_ - safety_margin)
               .order_by('-penalizado_lab_hasta', 'nombre'))

    return render(request, "impresoras/penalizados.html", {
        "imp_pen": imp_pen,
        "lab_pen": lab_pen,
    })
@login_required
def exportar_excel_lab(request):
    """
    Exporta EXCLUSIVAMENTE datos del laboratorio para el mes dado (?month=YYYY-MM).
    No depende de impresora_id ni del nombre del recurso.
    Crea 3 hojas: Detalle Lab, Asistencia Lab (Usados) y Resumen Lab.
    """
    from collections import defaultdict

    month_str = request.GET.get("month")  # "YYYY-MM"
    if not month_str:
        messages.error(request, "Selecciona un mes (YYYY-MM).")
        return redirect("impresoras:calendario")

    # Rango de fechas del mes
    try:
        y, m = map(int, month_str.split("-"))
        first = date(y, m, 1)
        last  = date(y, m, calendar.monthrange(y, m)[1])
    except Exception:
        messages.error(request, "Formato de mes inválido (usa YYYY-MM).")
        return redirect("impresoras:calendario")

    # ====== Query base del Lab (excluye cancelados) ======
    qs_lab = (LabReserva.objects
              .filter(fecha__range=(first, last))
              .exclude(estado="cancelado"))

    # Prepara filas (con columnas técnicas para ordenar)
    # Estructura de cada fila:
    # 0 Recurso | 1 Fecha | 2 HoraTxt | 3 EstadoDisp | 4 Nombre | 5 Cédula | 6 Celular | 7 Carrera
    # 8 Actividad | 9 Observaciones | 10 Creado | 11 hora_num (técnico) | 12 estado_raw (técnico)
    rows_lab = []
    for r in qs_lab:
        creado_local = timezone.localtime(r.creado_en).replace(tzinfo=None) if getattr(r, "creado_en", None) else None
        rows_lab.append((
            LAB_NAME, r.fecha,
            f"{r.hora:02d}:00 - {r.hora+1:02d}:00",
            r.get_estado_display() if hasattr(r, "get_estado_display") else (r.estado or "").title(),
            r.estudiante_nombre, r.estudiante_cedula, r.estudiante_celular, r.estudiante_carrera,
            (getattr(r, "actividad", "") or ""),  # 👈 NUEVO
            (getattr(r, "tecnico_observaciones", "") or "").strip().replace("\r\n", " ").replace("\n", " "),
            creado_local,
            r.hora,                               # idx 11: hora_num (técnico)
            (r.estado or "").lower(),             # idx 12: estado_raw (técnico)
        ))

    # ====== Crea workbook ======
    wb = Workbook()

    # ---- Hoja 1: Detalle Lab ----
    ws_lab_det = wb.active
    ws_lab_det.title = "Detalle Lab"

    hdr = ["Recurso", "Fecha", "Hora", "Estado", "Nombre", "Cédula", "Celular", "Carrera",
           "Actividad", "Observaciones", "Creado"]  # 👈 Actividad añadida
    ws_lab_det.append(hdr)

    estado_order = {"usado": 0, "reservado": 1}
    rows_lab_sorted = sorted(
        rows_lab,
        key=lambda x: (x[1], x[11], estado_order.get(x[12], 99), x[4] or "")
    )
    for r in rows_lab_sorted:
        ws_lab_det.append(r[:11])  # ahora son 11 columnas visibles

    # Formatos y estilos
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDDDDD")

    for c in ws_lab_det[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # formato de "Creado" -> columna K
    for cell in ws_lab_det["K"][1:]:
        if isinstance(cell.value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm"

    for row in ws_lab_det.iter_rows(min_row=1, max_row=ws_lab_det.max_row, min_col=1, max_col=len(hdr)):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col_cells in ws_lab_det.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws_lab_det.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)

    if ws_lab_det.max_row > 1:
        t2 = Table(displayName="TablaDetalleLab", ref=f"A1:K{ws_lab_det.max_row}")
        t2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws_lab_det.add_table(t2)
    else:
        ws_lab_det["A2"] = "Sin reservas de laboratorio en el mes."
        ws_lab_det.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(hdr))
        ws_lab_det["A2"].alignment = Alignment(horizontal="center")
        ws_lab_det["A2"].font = Font(italic=True, color="666666")

    # ---- Hoja 2: Asistencia Lab (Usados) ----
    ws_lab_used = wb.create_sheet("Asistencia Lab (Usados)")
    used_hdr = ["Fecha", "Hora", "Nombre", "Cédula", "Celular", "Carrera",
                "Actividad", "Observaciones", "Creado"]  # 👈 Actividad añadida
    ws_lab_used.append(used_hdr)

    rows_lab_used = [r for r in rows_lab if r[12] == "usado"]  # estado_raw ahora en 12
    rows_lab_used = sorted(rows_lab_used, key=lambda r: (r[1], r[11], r[4] or ""))  # hora_num en 11
    for r in rows_lab_used:
        ws_lab_used.append([
            r[1], r[2],              # Fecha, Hora (texto)
            r[4], r[5], r[6], r[7],  # Nombre, Cédula, Celular, Carrera
            r[8],                    # Actividad
            r[9],                    # Observaciones
            r[10],                   # Creado
        ])

    for c in ws_lab_used[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws_lab_used.iter_rows(min_row=1, max_row=ws_lab_used.max_row, min_col=1, max_col=len(used_hdr)):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col_cells in ws_lab_used.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws_lab_used.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)

    if ws_lab_used.max_row > 1:
        t4 = Table(displayName="TablaAsistenciaLab", ref=f"A1:I{ws_lab_used.max_row}")
        t4.tableStyleInfo = TableStyleInfo(name="TableStyleMedium7", showRowStripes=True)
        ws_lab_used.add_table(t4)
    else:
        ws_lab_used["A2"] = "Sin asistencias registradas (estado 'usado') en el mes."
        ws_lab_used.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(used_hdr))
        ws_lab_used["A2"].alignment = Alignment(horizontal="center")
        ws_lab_used["A2"].font = Font(italic=True, color="666666")

    # ---- Hoja 3: Resumen Lab ----
    resumen = defaultdict(lambda: {"total": 0, "usados": 0, "reservados": 0})
    for r in rows_lab:
        fecha = r[1]
        hora_num = r[11]     # antes 10
        estado = r[12]       # antes 11
        resumen[(fecha, hora_num)]["total"] += 1
        if estado == "usado":
            resumen[(fecha, hora_num)]["usados"] += 1
        elif estado == "reservado":
            resumen[(fecha, hora_num)]["reservados"] += 1

    ws_lab_sum = wb.create_sheet("Resumen Lab")
    sum_hdr = ["Fecha", "Hora", "Total", "Usados", "Reservados", "Cupo", "%Uso (Usados/Cupo)"]
    ws_lab_sum.append(sum_hdr)

    for (fecha, hora_num) in sorted(resumen.keys()):
        tot = resumen[(fecha, hora_num)]["total"]
        usados = resumen[(fecha, hora_num)]["usados"]
        reservados = resumen[(fecha, hora_num)]["reservados"]
        cupo = LAB_CAPACITY
        pct = usados / cupo if cupo else 0
        ws_lab_sum.append([
            fecha,
            f"{hora_num:02d}:00 - {hora_num+1:02d}:00",
            tot, usados, reservados, cupo, pct
        ])

    for c in ws_lab_sum[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws_lab_sum["G"][1:]:
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"
    for row in ws_lab_sum.iter_rows(min_row=1, max_row=ws_lab_sum.max_row, min_col=1, max_col=len(sum_hdr)):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for col_cells in ws_lab_sum.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws_lab_sum.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 32)

    if ws_lab_sum.max_row > 1:
        t3 = Table(displayName="TablaResumenLab", ref=f"A1:G{ws_lab_sum.max_row}")
        t3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        ws_lab_sum.add_table(t3)
    else:
        ws_lab_sum["A2"] = "Sin datos para el resumen de laboratorio."
        ws_lab_sum.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(sum_hdr))
        ws_lab_sum["A2"].alignment = Alignment(horizontal="center")
        ws_lab_sum["A2"].font = Font(italic=True, color="666666")

    # ====== Respuesta ======
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="laboratorio_{month_str}.xlsx"'
    return resp
