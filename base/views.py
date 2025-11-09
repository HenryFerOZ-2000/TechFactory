# Standard library
import csv
from io import BytesIO
from zoneinfo import ZoneInfo

# Third-party
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Django
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_http_methods, require_POST

# Local apps
from .models import Componente, Persona, Registro

# Constantes
ACTION_PASSWORD = getattr(settings, "ACTION_PASSWORD", "TechFactory2025")

def landing(request):
    return render(request, 'base/landing.html')


@require_POST
def componentes_delete(request, comp_id):
    comp = get_object_or_404(Componente, pk=comp_id)

    # Datos de contexto para el informe
    ahora = timezone.localtime(timezone.now())
    total_reg = Registro.objects.filter(componente=comp).count()
    abiertos = Registro.objects.filter(componente=comp, fecha_entrada__isnull=True).count()

    with transaction.atomic():
        try:
            # Intento de borrado físico
            comp.delete()
            accion = "Eliminado del inventario."
        except ProtectedError:
            # Tiene historial (FK protegida) → no se borra; se desactiva
            accion = "No se pudo eliminar por historial. Marcado como NO DISPONIBLE."
            fields = []
            if hasattr(comp, "activo"):
                comp.activo = False
                fields.append("activo")
            if hasattr(comp, "cantidad_disponible"):
                comp.cantidad_disponible = 0
                fields.append("cantidad_disponible")
            comp.save(update_fields=fields or None)

    # Informe .txt
    contenido = (
        "BAJA / NO DISPONIBLE - Tech Factory\n"
        f"Fecha: {ahora:%Y-%m-%d %H:%M}\n"
        "============================================================\n"
        f"Componente: {comp.nombre}\n"
        f"Ubicación : {comp.ubicacion}\n"
        f"Registros asociados: {total_reg}  (abiertos: {abiertos})\n"
        f"Acción: {accion}\n"
        "============================================================\n"
        "Este informe certifica que el componente ya no está disponible.\n"
    )

    resp = HttpResponse(contenido, content_type="text/plain; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="baja_{slugify(comp.nombre)}.txt"'
    return resp
def generar_correo(request, registro_id):
    """
    Descarga un .txt con el cuerpo de correo para el registro indicado.
    Si está vencido, se indica el atraso; si no, se deja como aviso de vencimiento.
    """
    r = get_object_or_404(
        Registro.objects.select_related('persona', 'componente'),
        pk=registro_id
    )

    now = timezone.now()
    vencido = (r.fecha_entrada is None) and (r.vence_el and r.vence_el <= now)
    dias_atraso = 0
    if vencido:
        dias_atraso = (now - r.vence_el).days

    fecha_salida = timezone.localtime(r.fecha_salida).strftime("%Y-%m-%d %H:%M")
    vence_el_txt = r.vence_el and timezone.localtime(r.vence_el).strftime("%Y-%m-%d %H:%M") or "—"
    fecha_entrada_txt = r.fecha_entrada and timezone.localtime(r.fecha_entrada).strftime("%Y-%m-%d %H:%M") or "—"

    asunto = f"Recordatorio de devolución: {r.componente.nombre}"
    saludo = f"Estimado/a {r.persona.nombre}:"
    cuerpo_vencido = (
        f"Se ha cumplido el plazo de devolución del componente solicitado.\n\n"
        f"• Componente: {r.componente.nombre}\n"
        f"• Cantidad: {r.cantidad}\n"
        f"• Ubicación: {r.componente.ubicacion}\n"
        f"• Fecha de salida: {fecha_salida}\n"
        f"• Fecha de vencimiento: {vence_el_txt}\n"
        f"• Días de atraso: {dias_atraso}\n\n"
        f"Le solicitamos acercarse a devolver el material o comunicarse para regularizar su situación."
    )
    cuerpo_aviso = (
        f"Le recordamos el próximo vencimiento del préstamo del componente solicitado.\n\n"
        f"• Componente: {r.componente.nombre}\n"
        f"• Cantidad: {r.cantidad}\n"
        f"• Ubicación: {r.componente.ubicacion}\n"
        f"• Fecha de salida: {fecha_salida}\n"
        f"• Fecha de vencimiento: {vence_el_txt}\n\n"
        f"Por favor, gestione la devolución en el tiempo establecido."
    )

    cierre = (
        "\n\nAtentamente,\n"
        "Tech Factory – Gestión de Inventario\n"
    )

    cuerpo = f"Asunto: {asunto}\n\n{saludo}\n\n" + (cuerpo_vencido if vencido else cuerpo_aviso) + cierre

    filename = f"correo_{r.id}_{slugify(r.persona.nombre)}.txt"
    resp = HttpResponse(cuerpo, content_type="text/plain; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def informe_vencidos(request):
    """
    Descarga un .txt con un listado de TODOS los registros vencidos (sin entrada).
    """
    now = timezone.now()
    vencidos = (
        Registro.objects
        .select_related('persona', 'componente')
        .filter(fecha_entrada__isnull=True, vence_el__lte=now)
        .order_by('persona__nombre', 'componente__nombre')
    )

    lineas = []
    lineas.append("INFORME DE VENCIDOS - Tech Factory\n")
    lineas.append(f"Generado: {timezone.localtime(now).strftime('%Y-%m-%d %H:%M')}\n")
    lineas.append("="*60 + "\n")

    if not vencidos.exists():
        lineas.append("No hay préstamos vencidos.\n")
    else:
        for r in vencidos:
            dias_atraso = (now - r.vence_el).days if r.vence_el else 0
            fecha_salida = timezone.localtime(r.fecha_salida).strftime("%Y-%m-%d %H:%M")
            vence_el_txt = r.vence_el and timezone.localtime(r.vence_el).strftime("%Y-%m-%d %H:%M") or "—"
            lineas.append(
                f"- Persona: {r.persona.nombre} ({r.persona.carrera}) | Cédula: {r.persona.cedula}\n"
                f"  Componente: {r.componente.nombre} | Cantidad: {r.cantidad} | Ubi: {r.componente.ubicacion}\n"
                f"  Salida: {fecha_salida} | Vence: {vence_el_txt} | Atraso: {dias_atraso} día(s)\n"
            )
            lineas.append("-"*60 + "\n")

    contenido = "".join(lineas)
    filename = "informe_vencidos.txt"
    resp = HttpResponse(contenido, content_type="text/plain; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

def eliminar_registro(request, registro_id):
    if request.method != 'POST':
        return redirect('base:index')

    password = (request.POST.get('password') or '').strip()
    if password != ACTION_PASSWORD:
        messages.error(request, 'Password incorrecta.')
        return redirect('base:index')

    r = get_object_or_404(Registro.objects.select_related('componente'), pk=registro_id)

    with transaction.atomic():
        # Si el registro NO fue devuelto, regresamos la cantidad al inventario
        if r.fecha_entrada is None and r.componente_id:
            comp = Componente.objects.select_for_update().get(pk=r.componente_id)
            comp.cantidad_disponible = (comp.cantidad_disponible or 0) + r.cantidad
            comp.save(update_fields=['cantidad_disponible'])

        # Luego eliminamos el registro
        r.delete()

    messages.success(request, 'Registro eliminado. (Stock ajustado si estaba prestado)')
    return redirect('base:index')


def limpiar_todo(request):
    if request.method != 'POST':
        return redirect('base:index')

    password = (request.POST.get('password') or '').strip()
    if password != ACTION_PASSWORD:
        messages.error(request, 'Password incorrecta.')
        return redirect('base:index')

    with transaction.atomic():
        # 1) Devolver al inventario SOLO lo que sigue prestado
        prestados = (
            Registro.objects
            .select_related('componente')
            .filter(fecha_entrada__isnull=True)
        )
        # Devolvemos por lote para evitar doble consulta por componente
        for r in prestados:
            if r.componente_id:
                comp = Componente.objects.select_for_update().get(pk=r.componente_id)
                comp.cantidad_disponible = (comp.cantidad_disponible or 0) + r.cantidad
                comp.save(update_fields=['cantidad_disponible'])

        # 2) Borrar todos los registros
        Registro.objects.all().delete()

    messages.success(request, 'Se limpiaron todos los registros. (Stock ajustado para los no devueltos)')
    return redirect('base:index')

def descargar_plantilla_csv(request):
    # Configurar la respuesta como un archivo CSV
    response = HttpResponse(
        content_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="plantilla_componentes.csv"'},
    )

    writer = csv.writer(response)
    # Cabeceras del CSV (ajusta a los campos de tu modelo)
    writer.writerow(["nombre", "ubicacion", "cantidad"])

    # Puedes poner filas de ejemplo (opcional)
    writer.writerow(["Resistor 10kΩ", "Caja A1", "50"])
    writer.writerow(["Arduino UNO", "Estante B2", "10"])

    return response


def importar_csv(request):
    if request.method == "POST":
        csv_file = request.FILES.get("archivo_csv")
        if not csv_file:
            messages.error(request, "Debes seleccionar un archivo CSV.")
            return redirect("base:componentes")

        try:
            # Decodificar y leer CSV
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)

            for row in reader:
                # Ajusta los nombres de columnas a tu CSV
                Componente.objects.create(
                    nombre=row["nombre"],
                    ubicacion=row["ubicacion"],
                    cantidad=row["cantidad"]
                )
            messages.success(request, "Datos importados correctamente.")
        except Exception as e:
            messages.error(request, f"Error al importar CSV: {e}")

        return redirect("base:componentes")

    return render(request, "base/importar_csv.html")
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        next_url = request.POST.get('next') or reverse('base:index')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)
            # Mensaje opcional
            messages.success(request, 'Acceso concedido.')
            return redirect(next_url)  # ← IMPORTANTE: REDIRIGE
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
            return render(request, 'base/login.html', {'next': next_url})

    # GET
    next_url = request.GET.get('next', reverse('base:index'))
    return render(request, 'base/login.html', {'next': next_url})

def logout_view(request):
    request.session.flush()
    return redirect("base:landing")


def exportar_excel(request):
    qs = (Registro.objects
          .select_related('persona', 'componente')
          .all())

    # Zona horaria local (ajusta en settings.TIME_ZONE)
    tzname = getattr(settings, "TIME_ZONE", "UTC")
    local_tz = ZoneInfo(tzname)

    def to_local_naive(dt):
        if not dt:
            return None
        if djtz.is_aware(dt):
            dt = djtz.localtime(dt, local_tz)
        return dt.replace(tzinfo=None)

    # Construimos las filas a mano para evitar FieldError
    rows = []
    for r in qs:
        rows.append({
            "Nombre": r.persona.nombre,
            "Cédula": str(r.persona.cedula or ""),          # texto para preservar ceros
            "Celular": str(r.persona.celular or ""),        # texto
            "Carrera": r.persona.carrera,
            "Componente": r.componente.nombre,
            "Ubicación": r.componente.ubicacion,
            "Cantidad": int(r.cantidad or 0),
            "Fecha/Hora Salida": to_local_naive(r.fecha_salida),
            "Fecha/Hora Entrada": to_local_naive(r.fecha_entrada),
            "Vence el": to_local_naive(r.vence_el),
            "Estado": r.estado,
            "Renovaciones": int(r.renovaciones or 0),
            "Zona horaria usada": tzname,
        })

    # Orden claro de columnas
    cols = [
        "Nombre","Cédula","Celular","Carrera","Componente","Ubicación","Cantidad",
        "Fecha/Hora Salida","Fecha/Hora Entrada","Vence el",
        "Estado","Renovaciones","Zona horaria usada"
    ]
    df = pd.DataFrame(rows, columns=cols)

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss") as writer:
        df.to_excel(writer, index=False, sheet_name="Registros")
        ws = writer.sheets["Registros"]

        # ===== Estilos =====
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        bold = Font(bold=True)
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        # Cabeceras
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Ajuste de columnas + formato de fecha
        for idx, col_name in enumerate(cols, start=1):
            width = max(len(col_name) + 2, 16)
            if col_name in ("Cédula", "Celular"):
                width = max(width, 18)
            if col_name in ("Fecha/Hora Salida", "Fecha/Hora Entrada", "Vence el"):
                width = max(width, 22)
                # aplicar formato explícito a las celdas de fecha
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=idx).number_format = "yyyy-mm-dd hh:mm:ss"

            # Bordes y alineación para toda la columna
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=idx)
                cell.border = border
                if row > 1:
                    cell.alignment = Alignment(vertical="center")

            ws.column_dimensions[get_column_letter(idx)].width = width

    out.seek(0)
    resp = HttpResponse(
        out.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="registros.xlsx"'
    return resp


@login_required
def index(request):
    registros = Registro.objects.select_related('persona','componente').all()

    # Filtros de la tabla
    fc   = request.GET.get('filtro_cedula','').strip()
    fcel = request.GET.get('filtro_celular','').strip()
    fcom = request.GET.get('filtro_componente','').strip()

    if fc:
        registros = registros.filter(persona__cedula__icontains=fc)
    if fcel:
        registros = registros.filter(persona__celular__icontains=fcel)
    if fcom:
        registros = registros.filter(componente__nombre__icontains=fcom)

    # Componentes para el datalist (también recortado por el mismo filtro)
    componentes = Componente.objects.all().order_by('nombre')
    if hasattr(Componente, 'activo'):
        componentes = componentes.filter(activo=True)
    if fcom:
        componentes = componentes.filter(nombre__icontains=fcom)

    # Opciones de carrera para el <select>
    class DummyForm: pass
    form = DummyForm()
    form.fields = {'carrera': type('X', (), {'choices': Persona.CARRERAS})}

    return render(request, 'base/index.html', {
        'form': form,
        'registros': registros,
        'componentes': componentes,
        'filtros': type('F', (), {'data': {'filtro_cedula': fc, 'filtro_celular': fcel, 'filtro_componente': fcom}})(),
        'now': timezone.now(),   # <-- NUEVO (aware)
    })




def registrar_salida(request):
    if request.method != 'POST':
        return redirect('base:index')

    nombre  = (request.POST.get('nombre')  or '').strip()
    cedula  = (request.POST.get('cedula')  or '').strip()
    celular = (request.POST.get('celular') or '').strip()
    carrera = (request.POST.get('carrera') or '').strip()
    
    # Validar nombre y apellido
    if nombre:
        partes = nombre.split()
        if len(partes) < 2:
            messages.error(request, 'Por favor ingrese nombre y apellido.')
            return redirect('base:index')
    
    # Validar que cédula solo contenga números
    if cedula and not cedula.isdigit():
        messages.error(request, 'La cédula solo puede contener números.')
        return redirect('base:index')
    
    # Validar que celular solo contenga números
    if celular and not celular.isdigit():
        messages.error(request, 'El celular solo puede contener números.')
        return redirect('base:index')

    # cantidad segura
    try:
        cantidad = int(request.POST.get('cantidad') or 0)
    except ValueError:
        cantidad = 0

    # Texto del datalist: puede ser "123 | Arduino Uno"
    # o "123 | Arduino Uno | UBI: Lab 1 | DISP: 5"
    combo_txt = (request.POST.get('componente_combo') or '').strip()

    comp_id = None
    comp_nom = None
    if '|' in combo_txt:
        parts = [p.strip() for p in combo_txt.split('|')]
        # parts[0] = id; parts[1] = nombre (las demás partes pueden ser UBI/DISP)
        if parts:
            comp_id = parts[0]
        if len(parts) >= 2:
            comp_nom = parts[1]
    else:
        comp_nom = combo_txt  # si alguien escribe a mano

    qs = Componente.objects.all()
    if hasattr(Componente, 'activo'):
        qs = qs.filter(activo=True)

    componente = None
    if comp_id and comp_id.isdigit():
        componente = qs.filter(id=int(comp_id)).first()

    # Fallback por nombre SOLO si no vino un ID válido
    if not componente and comp_nom:
        # exacto primero
        componente = qs.filter(nombre__iexact=comp_nom).first()
        if not componente:
            # si el usuario escribió algo parcial, intenta coincidencia única
            posibles = list(qs.filter(nombre__icontains=comp_nom)[:2])
            if len(posibles) == 1:
                componente = posibles[0]

    if not componente:
        messages.error(request, 'Selecciona un componente válido de la lista.')
        return redirect('base:index')

    if cantidad < 1:
        messages.error(request, 'Cantidad inválida.')
        return redirect('base:index')

    if (componente.cantidad_disponible or 0) < cantidad:
        messages.error(request, f'No hay suficientes unidades disponibles de {componente.nombre}.')
        return redirect('base:index')

    persona, _ = Persona.objects.get_or_create(
        cedula=cedula,
        defaults={'nombre': nombre, 'celular': celular, 'carrera': carrera}
    )
    # Actualiza datos por si cambiaron
    persona.nombre = nombre
    persona.celular = celular
    persona.carrera = carrera
    persona.save()

    Registro.objects.create(
        persona=persona,
        componente=componente,
        cantidad=cantidad,
        vence_el=timezone.now() + timezone.timedelta(days=7)
    )

    # descuenta stock
    componente.cantidad_disponible = (componente.cantidad_disponible or 0) - cantidad
    componente.save()

    messages.success(request, 'Salida registrada.')
    return redirect('base:index')






# --------- HOME (Préstamos) ---------


def registrar_entrada(request, registro_id):
    r = get_object_or_404(Registro, id=registro_id)
    if r.estado == 'devuelto':
        messages.info(request, 'Ese registro ya fue devuelto.')
        return redirect('base:index')
    r.estado = 'devuelto'
    r.fecha_entrada = timezone.now()
    r.save()
    comp = r.componente
    comp.cantidad_disponible += r.cantidad
    comp.save()
    messages.success(request, 'Entrada registrada y stock actualizado.')
    return redirect('base:index')

def renovar_salida(request, registro_id):
    r = get_object_or_404(Registro, id=registro_id)
    if r.estado == 'devuelto':
        messages.error(request, 'No se puede renovar un registro devuelto.')
        return redirect('base:index')
    r.renovaciones += 1
    r.vence_el = (r.vence_el or timezone.now()) + timezone.timedelta(days=7)
    r.save()
    messages.success(request, 'Salida renovada por 7 días.')
    return redirect('base:index')



# --------- Gestión de Componentes (para el botón "Gestionar componentes") ---------

@login_required
def componentes_list(request):
    q = request.GET.get("q", "").strip()
    comps = Componente.objects.all().order_by("nombre")
    if hasattr(Componente, 'activo') and request.GET.get('solo_activos') == '1':
        comps = comps.filter(activo=True)
    if q:
        comps = comps.filter(Q(nombre__icontains=q) | Q(ubicacion__icontains=q))
    return render(request, "base/componentes_list.html", {"comps": comps, "q": q})

@require_http_methods(["POST"])
def componentes_create(request):
    nombre = request.POST.get("nombre","").strip()
    ubic  = request.POST.get("ubicacion","").strip()
    total = int(request.POST.get("cantidad_total") or 0)
    disp  = int(request.POST.get("cantidad_disponible") or 0)
    activo = request.POST.get("activo") == "on"
    if not nombre:
        messages.error(request, "Nombre requerido.")
        return redirect("base:componentes_list")
    obj, created = Componente.objects.update_or_create(
        nombre=nombre,
        defaults={"ubicacion": ubic, "cantidad_total": total, "cantidad_disponible": disp, "activo": activo}
    )
    messages.success(request, "Componente creado/actualizado.")
    return redirect("base:componentes_list")

@require_http_methods(["POST"])
def componentes_edit(request, comp_id):
    comp = get_object_or_404(Componente, id=comp_id)
    nombre = request.POST.get("nombre","").strip()
    ubic  = request.POST.get("ubicacion","").strip()
    total = int(request.POST.get("cantidad_total") or 0)
    disp  = int(request.POST.get("cantidad_disponible") or 0)
    activo = request.POST.get("activo") == "on"
    
    if not nombre:
        messages.error(request, "Nombre requerido.")
        return redirect("base:componentes_list")
    
    comp.nombre = nombre
    comp.ubicacion = ubic
    comp.cantidad_total = total
    comp.cantidad_disponible = disp
    comp.activo = activo
    comp.save()
    
    messages.success(request, f"Componente '{nombre}' actualizado.")
    return redirect("base:componentes_list")

@require_http_methods(["POST"])
def componentes_toggle_activo(request, comp_id):
    comp = get_object_or_404(Componente, id=comp_id)
    comp.activo = not comp.activo
    comp.save()
    estado = "activado" if comp.activo else "desactivado"
    messages.success(request, f"Componente '{comp.nombre}' {estado}.")
    return redirect("base:componentes_list")


# --------- Cargar inventario (CSV) ---------
@require_http_methods(["GET","POST"])
def cargar_inventario(request):
    if request.method == "GET":
        return render(request, 'base/cargar_inventario.html', {})
    file: UploadedFile = request.FILES.get('csv')
    if not file or not file.name.lower().endswith('.csv'):
        messages.error(request, 'Sube un archivo CSV.')
        return redirect('base:cargar_inventario')
    decoded = file.read().decode('utf-8').splitlines()
    reader = csv.DictReader(decoded)
    creados, actualizados = 0, 0
    for row in reader:
        nombre = (row.get('nombre') or '').strip()
        if not nombre:
            continue
        defaults = {
            'ubicacion': (row.get('ubicacion') or '').strip(),
            'cantidad_total': int(row.get('cantidad_total') or 0),
            'cantidad_disponible': int(row.get('cantidad_disponible') or 0),
        }
        if 'activo' in row:
            defaults['activo'] = str(row.get('activo') or 'true').lower() in ('1','true','si','sí')
        obj, created = Componente.objects.update_or_create(nombre=nombre, defaults=defaults)
        creados += int(created)
        actualizados += int(not created)
    messages.success(request, f'Inventario cargado. Creados: {creados}, Actualizados: {actualizados}.')
    return redirect('base:componentes_list')
